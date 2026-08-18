from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.worksheet.datavalidation import (
    DataValidation,
)

from apps.academics.models import (
    AcademicYear,
    ClassSection,
)
from apps.guardians.models import (
    StudentGuardian,
)
from apps.students.models import Student


HEADERS = [
    "First Name",
    "Middle Name",
    "Last Name",
    "Date of Birth",
    "Gender",
    "Admission Date",
    "Phone Number",
    "Email",
    "Academic Year",
    "Class Level Code",
    "Class Section Code",
    "Enrolled On",
    "Guardian First Name",
    "Guardian Last Name",
    "Guardian Phone",
    "Guardian Email",
    "Guardian Relationship",
]


def _list_validation(
    values,
):
    cleaned = [
        str(value)
        for value in values
        if value
    ]

    value_string = ",".join(
        cleaned
    )

    if (
        not cleaned
        or len(value_string) > 250
    ):
        return None

    return DataValidation(
        type="list",
        formula1=(
            f'"{value_string}"'
        ),
        allow_blank=True,
    )


def build_student_import_template(
    *,
    school,
):
    workbook = Workbook()

    students_sheet = (
        workbook.active
    )

    students_sheet.title = (
        "Students"
    )

    instructions = (
        workbook.create_sheet(
            "Instructions"
        )
    )

    reference = (
        workbook.create_sheet(
            "Reference"
        )
    )

    # ---------------------------------------------------------
    # STUDENTS SHEET
    # ---------------------------------------------------------

    students_sheet.append(
        HEADERS
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="111827",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in students_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    students_sheet.freeze_panes = (
        "A2"
    )

    students_sheet.auto_filter.ref = (
        f"A1:Q5001"
    )

    widths = {
        "A": 18,
        "B": 18,
        "C": 18,
        "D": 15,
        "E": 15,
        "F": 15,
        "G": 17,
        "H": 28,
        "I": 18,
        "J": 20,
        "K": 22,
        "L": 15,
        "M": 22,
        "N": 22,
        "O": 18,
        "P": 28,
        "Q": 22,
    }

    for column, width in (
        widths.items()
    ):
        students_sheet.column_dimensions[
            column
        ].width = width

    for row in range(
        2,
        5002,
    ):
        students_sheet[
            f"D{row}"
        ].number_format = "yyyy-mm-dd"

        students_sheet[
            f"F{row}"
        ].number_format = "yyyy-mm-dd"

        students_sheet[
            f"L{row}"
        ].number_format = "yyyy-mm-dd"

    # ---------------------------------------------------------
    # CURRENT SCHOOL VALUES
    # ---------------------------------------------------------

    academic_years = list(
        AcademicYear.objects
        .for_school(school)
        .order_by(
            "-starts_on"
        )
        .values_list(
            "name",
            flat=True,
        )
    )

    sections = list(
        ClassSection.objects
        .for_school(school)
        .filter(
            is_active=True
        )
        .select_related(
            "level"
        )
        .order_by(
            "level__order",
            "name",
        )
    )

    level_codes = list(
        dict.fromkeys(
            section.level.code
            for section in sections
        )
    )

    section_codes = list(
        dict.fromkeys(
            section.code
            for section in sections
        )
    )

    gender_values = [
        choice
        for choice, _
        in Student.Gender.choices
    ]

    relationship_values = [
        choice
        for choice, _
        in (
            StudentGuardian
            .Relationship
            .choices
        )
    ]

    validations = [
        (
            "E2:E5001",
            gender_values,
        ),
        (
            "I2:I5001",
            academic_years,
        ),
        (
            "J2:J5001",
            level_codes,
        ),
        (
            "K2:K5001",
            section_codes,
        ),
        (
            "Q2:Q5001",
            relationship_values,
        ),
    ]

    for cell_range, values in validations:

        validation = (
            _list_validation(
                values
            )
        )

        if validation:
            students_sheet.add_data_validation(
                validation
            )

            validation.add(
                cell_range
            )

    # ---------------------------------------------------------
    # INSTRUCTIONS
    # ---------------------------------------------------------

    instructions.column_dimensions[
        "A"
    ].width = 28

    instructions.column_dimensions[
        "B"
    ].width = 75

    instruction_rows = [
        [
            "Student Bulk Import",
            (
                "Complete the Students sheet "
                "and upload the finished XLSX file."
            ),
        ],
        [
            "Admission Number",
            (
                "Do not enter an admission number. "
                "The system generates it automatically."
            ),
        ],
        [
            "Required Fields",
            (
                "First Name, Last Name, Academic Year, "
                "Class Level Code, Class Section Code, "
                "and Enrolled On."
            ),
        ],
        [
            "Dates",
            "Use YYYY-MM-DD.",
        ],
        [
            "Academic Year",
            (
                "Use one of the academic year values "
                "shown on the Reference sheet."
            ),
        ],
        [
            "Class",
            (
                "Use a valid Class Level Code and "
                "Class Section Code combination from "
                "the Reference sheet."
            ),
        ],
        [
            "Guardian",
            (
                "Guardian details are optional. "
                "If provided, include first name, "
                "last name and phone number."
            ),
        ],
        [
            "Headers",
            (
                "Do not rename, delete or move "
                "the column headers."
            ),
        ],
    ]

    for row in instruction_rows:
        instructions.append(
            row
        )

    instructions[
        "A1"
    ].font = Font(
        bold=True
    )

    # ---------------------------------------------------------
    # REFERENCE SHEET
    # ---------------------------------------------------------

    reference.append([
        "Academic Year",
        "Class Level",
        "Class Level Code",
        "Class Section",
        "Class Section Code",
    ])

    reference_row = 2

    maximum_rows = max(
        len(academic_years),
        len(sections),
        1,
    )

    for index in range(
        maximum_rows
    ):
        academic_year = (
            academic_years[index]
            if index < len(
                academic_years
            )
            else ""
        )

        section = (
            sections[index]
            if index < len(
                sections
            )
            else None
        )

        reference.cell(
            row=reference_row,
            column=1,
            value=academic_year,
        )

        if section:
            reference.cell(
                row=reference_row,
                column=2,
                value=(
                    section.level.name
                ),
            )

            reference.cell(
                row=reference_row,
                column=3,
                value=(
                    section.level.code
                ),
            )

            reference.cell(
                row=reference_row,
                column=4,
                value=(
                    section.name
                ),
            )

            reference.cell(
                row=reference_row,
                column=5,
                value=(
                    section.code
                ),
            )

        reference_row += 1

    for cell in reference[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column in [
        "A",
        "B",
        "C",
        "D",
        "E",
    ]:
        reference.column_dimensions[
            column
        ].width = 24

    output = BytesIO()

    workbook.save(
        output
    )

    return output.getvalue()