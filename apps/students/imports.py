import csv
import io

from datetime import (
    date,
    datetime,
)

from pathlib import Path

from django.core.exceptions import (
    ValidationError,
)

from django.db import transaction

from django.utils import timezone

from openpyxl import load_workbook


from apps.academics.models import (
    AcademicYear,
    ClassLevel,
    ClassSection,
)

from apps.academics.services import (
    enroll_student,
)

from apps.audit.models import (
    AuditEvent,
)

from apps.guardians.models import (
    Guardian,
    StudentGuardian,
)

from apps.guardians.services import (
    link_guardian_to_student,
)

from apps.subscriptions.services import (
    assert_can_add_active_students,
)


from .models import (
    Student,
    StudentImportBatch,
    StudentImportRow,
)

from .services import (
    generate_admission_number,
)


MAX_IMPORT_ROWS = 5000


# Admission number is intentionally NOT here.
# It is generated automatically by the system.

REQUIRED_COLUMNS = {
    "first_name",
    "last_name",
    "academic_year",
    "class_level_code",
    "class_section_code",
    "enrolled_on",
}


# ============================================================
# NORMALIZATION
# ============================================================


def normalize_header(
    value,
):
    return (
        str(
            value or ""
        )
        .strip()
        .lower()
        .replace(
            " ",
            "_",
        )
        .replace(
            "-",
            "_",
        )
    )


def normalize_value(
    value,
):
    if value is None:
        return ""

    if isinstance(
        value,
        datetime,
    ):
        return (
            value.date()
            .isoformat()
        )

    if isinstance(
        value,
        date,
    ):
        return (
            value.isoformat()
        )

    return (
        str(
            value
        )
        .strip()
    )


# ============================================================
# DATE PARSING
# ============================================================


def parse_date_value(
    value,
):
    if not value:
        return None

    if isinstance(
        value,
        datetime,
    ):
        return (
            value.date()
        )

    if isinstance(
        value,
        date,
    ):
        return value

    text = (
        str(
            value
        )
        .strip()
    )

    formats = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y",
    ]

    for fmt in formats:

        try:
            return (
                datetime.strptime(
                    text,
                    fmt,
                )
                .date()
            )

        except ValueError:
            continue

    raise ValidationError(
        (
            f"Invalid date '{text}'. "
            "Use YYYY-MM-DD."
        )
    )


# ============================================================
# CSV
# ============================================================


def read_csv_file(
    uploaded_file,
):
    raw = (
        uploaded_file.read()
    )

    try:
        text = raw.decode(
            "utf-8-sig"
        )

    except UnicodeDecodeError as exc:

        raise ValidationError(
            (
                "CSV must use UTF-8 "
                "encoding."
            )
        ) from exc

    reader = csv.DictReader(
        io.StringIO(
            text
        )
    )

    if not reader.fieldnames:
        raise ValidationError(
            (
                "CSV has no header "
                "row."
            )
        )

    headers = [
        normalize_header(
            header
        )
        for header
        in reader.fieldnames
    ]

    if len(
        headers
    ) != len(
        set(
            headers
        )
    ):
        raise ValidationError(
            (
                "The import file contains "
                "duplicate column headers."
            )
        )

    rows = []

    for values in reader:

        # Ignore completely blank rows.
        if not any(
            normalize_value(
                value
            )
            for value
            in values.values()
        ):
            continue

        normalized = {}

        for (
            original_header,
            normalized_header,
        ) in zip(
            reader.fieldnames,
            headers,
        ):
            normalized[
                normalized_header
            ] = normalize_value(
                values.get(
                    original_header
                )
            )

        rows.append(
            normalized
        )

    return (
        headers,
        rows,
    )


# ============================================================
# XLSX
# ============================================================


def read_xlsx_file(
    uploaded_file,
):
    try:

        workbook = load_workbook(
            uploaded_file,
            read_only=True,
            data_only=True,
        )

    except Exception as exc:

        raise ValidationError(
            (
                "Unable to read "
                "XLSX file."
            )
        ) from exc

    sheet = (
        workbook.active
    )

    iterator = (
        sheet.iter_rows(
            values_only=True
        )
    )

    try:
        first_row = next(
            iterator
        )

    except StopIteration as exc:

        raise ValidationError(
            (
                "Spreadsheet is "
                "empty."
            )
        ) from exc

    headers = [
        normalize_header(
            value
        )
        for value
        in first_row
    ]

    # Remove trailing completely blank headers.
    while (
        headers
        and not headers[-1]
    ):
        headers.pop()

    if not headers:
        raise ValidationError(
            (
                "Spreadsheet has no "
                "header row."
            )
        )

    if len(
        headers
    ) != len(
        set(
            headers
        )
    ):
        raise ValidationError(
            (
                "The import file contains "
                "duplicate column headers."
            )
        )

    rows = []

    for row in iterator:

        relevant_row = (
            row[
                :len(
                    headers
                )
            ]
        )

        if not any(
            value is not None
            and str(
                value
            ).strip()
            for value
            in relevant_row
        ):
            continue

        normalized = {}

        for (
            index,
            header,
        ) in enumerate(
            headers
        ):

            if not header:
                continue

            value = (
                relevant_row[
                    index
                ]
                if index
                < len(
                    relevant_row
                )
                else ""
            )

            normalized[
                header
            ] = normalize_value(
                value
            )

        rows.append(
            normalized
        )

    return (
        headers,
        rows,
    )


# ============================================================
# FILE READER
# ============================================================


def read_import_file(
    uploaded_file,
):
    extension = (
        Path(
            uploaded_file.name
        )
        .suffix
        .lower()
    )

    if (
        extension
        == ".csv"
    ):
        return read_csv_file(
            uploaded_file
        )

    if (
        extension
        == ".xlsx"
    ):
        return read_xlsx_file(
            uploaded_file
        )

    raise ValidationError(
        (
            "Unsupported file "
            "type."
        )
    )


# ============================================================
# ROW VALIDATION
# ============================================================


def validate_import_row(
    *,
    school,
    row,
):
    errors = []

    normalized = {
        key: normalize_value(
            value
        )
        for key, value
        in row.items()
    }

    # --------------------------------------------------------
    # Student name
    # --------------------------------------------------------

    first_name = (
        normalized.get(
            "first_name",
            ""
        )
    )

    last_name = (
        normalized.get(
            "last_name",
            ""
        )
    )

    if not first_name:
        errors.append(
            (
                "First name is "
                "required."
            )
        )

    if not last_name:
        errors.append(
            (
                "Last name is "
                "required."
            )
        )

    # --------------------------------------------------------
    # Academic year
    # --------------------------------------------------------

    academic_year_name = (
        normalized.get(
            "academic_year",
            ""
        )
    )

    academic_year = (
        AcademicYear.objects
        .for_school(
            school
        )
        .filter(
            name=(
                academic_year_name
            )
        )
        .first()
    )

    if not academic_year:
        errors.append(
            (
                "Academic year "
                f"'{academic_year_name}' "
                "was not found."
            )
        )

    # --------------------------------------------------------
    # Class level
    # --------------------------------------------------------

    level_code = (
        normalized.get(
            "class_level_code",
            ""
        )
    )

    class_level = (
        ClassLevel.objects
        .for_school(
            school
        )
        .filter(
            code=level_code
        )
        .first()
    )

    if not class_level:
        errors.append(
            (
                "Class level code "
                f"'{level_code}' "
                "was not found."
            )
        )

    # --------------------------------------------------------
    # Class section
    # --------------------------------------------------------

    section_code = (
        normalized.get(
            "class_section_code",
            ""
        )
    )

    class_section = None

    if class_level:

        class_section = (
            ClassSection.objects
            .for_school(
                school
            )
            .filter(
                level=(
                    class_level
                ),
                code=(
                    section_code
                ),
                is_active=True,
            )
            .first()
        )

    if not class_section:
        errors.append(
            (
                "Class section "
                f"'{section_code}' "
                "was not found for "
                f"'{level_code}'."
            )
        )

    # --------------------------------------------------------
    # Enrolled on
    # --------------------------------------------------------

    try:
        enrolled_on = (
            parse_date_value(
                normalized.get(
                    "enrolled_on"
                )
            )
        )

    except ValidationError as exc:

        enrolled_on = None

        errors.extend(
            exc.messages
        )

    if not enrolled_on:
        errors.append(
            (
                "Enrollment date "
                "is required."
            )
        )

    if (
        academic_year
        and enrolled_on
        and not (
            academic_year.starts_on
            <= enrolled_on
            <= academic_year.ends_on
        )
    ):
        errors.append(
            (
                "Enrollment date is "
                "outside the academic "
                "year."
            )
        )

    # --------------------------------------------------------
    # Admission date
    # --------------------------------------------------------

    try:
        admission_date = (
            parse_date_value(
                normalized.get(
                    "admission_date"
                )
            )
        )

    except ValidationError as exc:

        admission_date = None

        errors.extend(
            exc.messages
        )

    # If admission date is blank,
    # use enrollment date.
    if (
        admission_date is None
        and enrolled_on
    ):
        admission_date = (
            enrolled_on
        )

    if (
        admission_date
        and enrolled_on
        and admission_date
        > enrolled_on
    ):
        errors.append(
            (
                "Admission date cannot "
                "be after enrollment "
                "date."
            )
        )

    # --------------------------------------------------------
    # Date of birth
    # --------------------------------------------------------

    try:
        date_of_birth = (
            parse_date_value(
                normalized.get(
                    "date_of_birth"
                )
            )
        )

    except ValidationError as exc:

        date_of_birth = None

        errors.extend(
            exc.messages
        )

    if (
        date_of_birth
        and admission_date
        and date_of_birth
        >= admission_date
    ):
        errors.append(
            (
                "Date of birth must "
                "be before admission "
                "date."
            )
        )

    # --------------------------------------------------------
    # Gender
    # --------------------------------------------------------

    gender = (
        normalized.get(
            "gender"
        )
        or Student
        .Gender
        .UNSPECIFIED
    ).lower()

    valid_genders = {
        choice
        for choice, _
        in Student
        .Gender
        .choices
    }

    if (
        gender
        not in valid_genders
    ):
        errors.append(
            (
                "Invalid gender "
                f"'{gender}'."
            )
        )

    # --------------------------------------------------------
    # Guardian
    # --------------------------------------------------------

    guardian_first_name = (
        normalized.get(
            "guardian_first_name",
            ""
        )
    )

    guardian_last_name = (
        normalized.get(
            "guardian_last_name",
            ""
        )
    )

    guardian_phone = (
        normalized.get(
            "guardian_phone",
            ""
        )
    )

    guardian_email = (
        normalized.get(
            "guardian_email",
            ""
        )
    )

    relationship = (
        normalized.get(
            "guardian_relationship",
            ""
        )
        .lower()
    )

    guardian_values = [
        guardian_first_name,
        guardian_last_name,
        guardian_phone,
        guardian_email,
        relationship,
    ]

    if any(
        guardian_values
    ):

        if not (
            guardian_first_name
            and guardian_last_name
            and guardian_phone
        ):
            errors.append(
                (
                    "When guardian details "
                    "are provided, guardian "
                    "first name, last name "
                    "and phone number are "
                    "required."
                )
            )

    valid_relationships = {
        choice
        for choice, _
        in (
            StudentGuardian
            .Relationship
            .choices
        )
    }

    if (
        relationship
        and relationship
        not in valid_relationships
    ):
        errors.append(
            (
                "Invalid guardian "
                "relationship "
                f"'{relationship}'."
            )
        )

    # --------------------------------------------------------
    # Store resolved values
    # --------------------------------------------------------

    normalized.update(
        {
            "date_of_birth": (
                date_of_birth
                .isoformat()
                if date_of_birth
                else None
            ),

            "admission_date": (
                admission_date
                .isoformat()
                if admission_date
                else None
            ),

            "enrolled_on": (
                enrolled_on
                .isoformat()
                if enrolled_on
                else None
            ),

            "gender":
                gender,

            "guardian_relationship":
                relationship,

            "academic_year_id": (
                str(
                    academic_year.id
                )
                if academic_year
                else None
            ),

            "class_section_id": (
                str(
                    class_section.id
                )
                if class_section
                else None
            ),
        }
    )

    return (
        normalized,
        errors,
    )


# ============================================================
# STAGE IMPORT
# ============================================================


@transaction.atomic
def stage_student_import(
    *,
    school,
    uploaded_file,
    uploaded_by,
):
    headers, rows = (
        read_import_file(
            uploaded_file
        )
    )

    missing = (
        REQUIRED_COLUMNS
        - set(
            headers
        )
    )

    if missing:

        raise ValidationError(
            (
                "Missing required "
                "columns: "
                + ", ".join(
                    sorted(
                        missing
                    )
                )
            )
        )

    if not rows:

        raise ValidationError(
            (
                "The import file "
                "contains no student "
                "rows."
            )
        )

    if (
        len(
            rows
        )
        > MAX_IMPORT_ROWS
    ):
        raise ValidationError(
            (
                "Import cannot exceed "
                f"{MAX_IMPORT_ROWS} "
                "rows."
            )
        )

    batch = (
        StudentImportBatch
        .objects
        .create(
            school=school,

            original_filename=(
                uploaded_file.name
            ),

            uploaded_by=(
                uploaded_by
            ),

            total_rows=(
                len(
                    rows
                )
            ),
        )
    )

    valid_count = 0

    invalid_count = 0

    for (
        index,
        row,
    ) in enumerate(
        rows,
        start=2,
    ):

        normalized, errors = (
            validate_import_row(
                school=school,
                row=row,
            )
        )

        is_valid = (
            not errors
        )

        if is_valid:
            valid_count += 1

        else:
            invalid_count += 1

        (
            StudentImportRow
            .objects
            .create(
                school=school,

                batch=batch,

                row_number=index,

                raw_data=row,

                normalized_data=(
                    normalized
                ),

                errors=errors,

                is_valid=(
                    is_valid
                ),
            )
        )

    batch.valid_rows = (
        valid_count
    )

    batch.invalid_rows = (
        invalid_count
    )

    batch.save(
        update_fields=[
            "valid_rows",
            "invalid_rows",
            "updated_at",
        ]
    )

    return batch


# ============================================================
# CONFIRM IMPORT
# ============================================================


@transaction.atomic
def confirm_student_import(
    *,
    batch,
    confirmed_by,
):
    batch = (
        StudentImportBatch.objects
        .select_for_update()
        .get(
            pk=batch.pk
        )
    )

    if (
        batch.status
        != StudentImportBatch
        .Status
        .STAGED
    ):
        raise ValidationError(
            (
                "Only staged imports "
                "can be confirmed."
            )
        )

    if (
        batch.invalid_rows
        > 0
    ):
        raise ValidationError(
            (
                "Fix all invalid rows "
                "before importing."
            )
        )

    rows = list(
        batch.rows
        .filter(
            is_valid=True
        )
        .order_by(
            "row_number"
        )
    )

    if not rows:
        raise ValidationError(
            (
                "There are no valid "
                "students to import."
            )
        )

    # --------------------------------------------------------
    # Subscription capacity
    # --------------------------------------------------------

    assert_can_add_active_students(
        school=(
            batch.school
        ),

        additional_count=(
            len(
                rows
            )
        ),
    )

    # --------------------------------------------------------
    # Processing status
    # --------------------------------------------------------

    batch.status = (
        StudentImportBatch
        .Status
        .PROCESSING
    )

    batch.save(
        update_fields=[
            "status",
            "updated_at",
        ]
    )

    imported = []

    # --------------------------------------------------------
    # Import each student
    # --------------------------------------------------------

    for row in rows:

        data = (
            row.normalized_data
        )

        # ----------------------------------------------------
        # Resolve academic year
        # ----------------------------------------------------

        academic_year = (
            AcademicYear.objects
            .for_school(
                batch.school
            )
            .get(
                id=(
                    data[
                        "academic_year_id"
                    ]
                )
            )
        )

        # ----------------------------------------------------
        # Resolve class
        # ----------------------------------------------------

        class_section = (
            ClassSection.objects
            .for_school(
                batch.school
            )
            .get(
                id=(
                    data[
                        "class_section_id"
                    ]
                )
            )
        )

        # ----------------------------------------------------
        # Dates
        # ----------------------------------------------------

        enrolled_on = (
            parse_date_value(
                data.get(
                    "enrolled_on"
                )
            )
        )

        admission_date = (
            parse_date_value(
                data.get(
                    "admission_date"
                )
            )
        )

        if (
            admission_date is None
        ):
            admission_date = (
                enrolled_on
            )

        # ----------------------------------------------------
        # Generate admission number
        # ----------------------------------------------------

        admission_number = (
            generate_admission_number(
                school=(
                    batch.school
                ),

                admission_date=(
                    admission_date
                ),

                academic_year=(
                    academic_year
                ),
            )
        )

        # ----------------------------------------------------
        # Create student
        # ----------------------------------------------------

        student = Student(
            school=(
                batch.school
            ),

            admission_number=(
                admission_number
            ),

            first_name=(
                data[
                    "first_name"
                ]
            ),

            middle_name=(
                data.get(
                    "middle_name",
                    "",
                )
            ),

            last_name=(
                data[
                    "last_name"
                ]
            ),

            date_of_birth=(
                parse_date_value(
                    data.get(
                        "date_of_birth"
                    )
                )
            ),

            gender=(
                data[
                    "gender"
                ]
            ),

            admission_date=(
                admission_date
            ),

            phone_number=(
                data.get(
                    "phone_number",
                    "",
                )
            ),

            email=(
                data.get(
                    "email",
                    "",
                )
            ),

            status=(
                Student
                .Status
                .ACTIVE
            ),
        )

        student.full_clean()

        student.save()

        # ----------------------------------------------------
        # Enrollment
        # ----------------------------------------------------

        enroll_student(
            school=(
                batch.school
            ),

            student=(
                student
            ),

            academic_year=(
                academic_year
            ),

            class_section=(
                class_section
            ),

            enrolled_on=(
                enrolled_on
            ),
        )

        # ----------------------------------------------------
        # Guardian
        # ----------------------------------------------------

        guardian_phone = (
            data.get(
                "guardian_phone",
                ""
            )
        )

        if guardian_phone:

            guardian = (
                Guardian.objects
                .for_school(
                    batch.school
                )
                .filter(
                    first_name__iexact=(
                        data.get(
                            "guardian_first_name",
                            "",
                        )
                    ),

                    last_name__iexact=(
                        data.get(
                            "guardian_last_name",
                            "",
                        )
                    ),

                    phone_number=(
                        guardian_phone
                    ),
                )
                .first()
            )

            if not guardian:

                guardian = Guardian(
                    school=(
                        batch.school
                    ),

                    first_name=(
                        data.get(
                            "guardian_first_name",
                            "",
                        )
                    ),

                    last_name=(
                        data.get(
                            "guardian_last_name",
                            "",
                        )
                    ),

                    phone_number=(
                        guardian_phone
                    ),

                    email=(
                        data.get(
                            "guardian_email",
                            "",
                        )
                    ),
                )

                guardian.full_clean()

                guardian.save()

            relationship = (
                data.get(
                    "guardian_relationship"
                )
                or (
                    StudentGuardian
                    .Relationship
                    .GUARDIAN
                )
            )

            link_guardian_to_student(
                school=(
                    batch.school
                ),

                student=(
                    student
                ),

                guardian=(
                    guardian
                ),

                relationship=(
                    relationship
                ),

                is_primary_contact=True,

                receives_reports=True,
            )

        # ----------------------------------------------------
        # Link staged row to imported student
        # ----------------------------------------------------

        row.imported_student = (
            student
        )

        row.save(
            update_fields=[
                "imported_student",
                "updated_at",
            ]
        )

        imported.append(
            student
        )

    # --------------------------------------------------------
    # Complete batch
    # --------------------------------------------------------

    batch.status = (
        StudentImportBatch
        .Status
        .COMPLETED
    )

    batch.completed_at = (
        timezone.now()
    )

    batch.save(
        update_fields=[
            "status",
            "completed_at",
            "updated_at",
        ]
    )

    # --------------------------------------------------------
    # Audit
    # --------------------------------------------------------

    AuditEvent.objects.create(
        school=(
            batch.school
        ),

        actor=(
            confirmed_by
        ),

        action=(
            "student_bulk_import_completed"
        ),

        object_type=(
            "StudentImportBatch"
        ),

        object_id=(
            str(
                batch.id
            )
        ),

        changes={
            "filename": (
                batch.original_filename
            ),

            "students_imported": (
                len(
                    imported
                )
            ),
        },
    )

    return imported